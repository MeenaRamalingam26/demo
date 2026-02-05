# from flask import (
#     Flask, render_template, request, redirect, url_for,
#     session, send_from_directory, abort
# )
# import sqlite3, os, sys
# from werkzeug.security import generate_password_hash, check_password_hash

# app = Flask(__name__)
# app.secret_key = "change-this-to-a-strong-secret"

# # Paths
# BASE_DIR = os.path.abspath(os.path.dirname(__file__))
# DB_PATH = os.path.join(BASE_DIR, "users.db")
# STATIC_DIR = os.path.join(BASE_DIR, "static")
# SAMPLES_DIR = os.path.join(STATIC_DIR, "samples")
# os.makedirs(SAMPLES_DIR, exist_ok=True)

# def get_conn():
#     return sqlite3.connect(DB_PATH)

# def init_db():
#     conn = get_conn()
#     c = conn.cursor()
#     c.execute('''CREATE TABLE IF NOT EXISTS users (
#                     id INTEGER PRIMARY KEY AUTOINCREMENT,
#                     username TEXT UNIQUE NOT NULL,
#                     password TEXT NOT NULL
#                 )''')
#     conn.commit()
#     conn.close()

# init_db()

# # --------- Try importing openpyxl once at startup for clear diagnostics ----------
# OPENPYXL_IMPORT_ERROR = None
# LOAD_WORKBOOK = None
# try:
#     from openpyxl import load_workbook  # type: ignore
#     LOAD_WORKBOOK = load_workbook
# except Exception as e:
#     OPENPYXL_IMPORT_ERROR = repr(e)

# # ------------------- Auth & Core -------------------
# @app.route('/')
# def home():
#     if session.get("username"):
#         return redirect(url_for('claims'))
#     return render_template('login.html')

# @app.route('/s3display')
# def s3display():
#     return render_template('s3display.html')

# @app.route('/login', methods=['POST'])
# def login():
#     username = request.form.get('username', '').strip()
#     password = request.form.get('password', '')

#     if not username or not password:
#         return render_template('login.html', error="Username and password are required"), 400

#     conn = get_conn()
#     c = conn.cursor()
#     c.execute("SELECT password FROM users WHERE username = ?", (username,))
#     row = c.fetchone()
#     conn.close()

#     if row is None:
#         return render_template('login.html', error="Invalid username or password"), 400

#     stored_hash = row[0]
#     if check_password_hash(stored_hash, password):
#         session['username'] = username
#         return redirect(url_for('claims'))
#     else:
#         return render_template('login.html', error="Invalid username or password"), 400

# @app.route('/logout')
# def logout():
#     session.clear()
#     return redirect(url_for('home'))

# # ------------------- UI pages -------------------
# @app.route('/claims')
# def claims():
#     return render_template('claims.html')

# # ✅ FIXED: use real Flask dynamic segment, not HTML-escaped
# @app.route('/review/<claim_id>')
# def review_claim(claim_id):
#     return render_template('review.html', claim_id=claim_id)

# # ------------------- Excel Preview (openpyxl-only) + Download -------------------
# def ws_to_html(ws, max_rows=None, max_cols=None):
#     """
#     Convert an openpyxl worksheet to a simple HTML table string.
#     - max_rows/max_cols: optional limits for very large sheets (None = no limit)
#     """
#     from openpyxl.utils import get_column_letter
#     from datetime import datetime, date, time

#     def fmt(val):
#         if isinstance(val, (datetime, date, time)):
#             try:
#                 # Safe iso-like formatting
#                 return val.isoformat(sep=' ')
#             except TypeError:
#                 return str(val)
#         return "" if val is None else str(val)

#     # Determine bounds
#     max_row = ws.max_row if not max_rows else min(ws.max_row, max_rows)
#     max_col = ws.max_column if not max_cols else min(ws.max_column, max_cols)

#     # Header row (A, B, C...)
#     thead_cells = [f"<th>{get_column_letter(c)}</th>" for c in range(1, max_col + 1)]
#     thead_html = "<thead><tr>" + "".join(thead_cells) + "</tr></thead>"

#     # Body rows
#     body_rows = []
#     for r in range(1, max_row + 1):
#         tds = []
#         for c in range(1, max_col + 1):
#             cell = ws.cell(row=r, column=c)
#             tds.append(f"<td>{fmt(cell.value)}</td>")
#         body_rows.append("<tr>" + "".join(tds) + "</tr>")
#     tbody_html = "<tbody>" + "".join(body_rows) + "</tbody>"

#     return f'<table class="excel-table">{thead_html}{tbody_html}</table>'

# @app.route('/excel-preview/<claim_id>')
# def excel_preview(claim_id):
#     """
#     Server-side preview of static/samples/dummy.xlsx using openpyxl only.
#     Renders each sheet as an HTML table (no client-side libs).
#     """
#     # If import failed at startup, show friendly diagnostics
#     if LOAD_WORKBOOK is None:
#         details = (
#             f"Import error: {OPENPYXL_IMPORT_ERROR}\n"
#             f"Python: {sys.version}\n"
#             f"Executable: {sys.executable}\n"
#             f"Sys.path[0]: {sys.path[0]}"
#         )
#         return render_template(
#             'excel_preview.html',
#             claim_id=claim_id,
#             sheets=[],
#             error="openpyxl import failed. See details below.",
#             diag=details
#         )

#     sample_filename = 'dummy.xlsx'
#     sample_abs_path = os.path.join(SAMPLES_DIR, sample_filename)
#     if not os.path.exists(sample_abs_path):
#         return render_template(
#             'excel_preview.html',
#             claim_id=claim_id,
#             sheets=[],
#             error="Dummy Excel not found at static/samples/dummy.xlsx",
#             diag=None
#         )

#     try:
#         # read_only=True for performance and to avoid locking
#         wb = LOAD_WORKBOOK(sample_abs_path, data_only=True, read_only=True)
#         sheets = []
#         # Optional: set max_rows/max_cols if needed for huge workbooks
#         for ws in wb.worksheets:
#             html_table = ws_to_html(ws)  # ws_to_html(ws, max_rows=1000, max_cols=50)
#             sheets.append({"name": ws.title, "table_html": html_table})
#     except Exception as e:
#         # Show the exact failure cause in the UI
#         return render_template(
#             'excel_preview.html',
#             claim_id=claim_id,
#             sheets=[],
#             error=f"Failed to read dummy.xlsx: {e}",
#             diag=f"Python: {sys.version}\nExecutable: {sys.executable}"
#         )

#     return render_template('excel_preview.html', claim_id=claim_id, sheets=sheets, error=None, diag=None)

# @app.route('/excel/download')
# def excel_download():
#     """
#     Downloads the dummy.xlsx file from static/samples/.
#     """
#     filename = 'dummy.xlsx'
#     full_path = os.path.join(SAMPLES_DIR, filename)
#     if not os.path.exists(full_path):
#         abort(404, description="Dummy Excel not found at static/samples/dummy.xlsx")
#     return send_from_directory(SAMPLES_DIR, filename, as_attachment=True)

# # -------- Optional: quick diagnostics route --------
# @app.route('/diag')
# def diag():
#     return {
#         "python_version": sys.version,
#         "executable": sys.executable,
#         "openpyxl_import_error": OPENPYXL_IMPORT_ERROR,
#         "cwd": os.getcwd(),
#         "base_dir": BASE_DIR,
#         "static_dir": STATIC_DIR,
#         "samples_dir_exists": os.path.isdir(SAMPLES_DIR),
#         "dummy_exists": os.path.exists(os.path.join(SAMPLES_DIR, "dummy.xlsx")),
#         "sys_path_0": sys.path[0],
#     }

# if __name__ == '__main__':
#     app.run(debug=True)
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, send_from_directory, abort
)
import sqlite3, os, sys
from werkzeug.security import check_password_hash

app = Flask(__name__)
app.secret_key = "change-this-to-a-strong-secret"

# Paths
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "users.db")
STATIC_DIR = os.path.join(BASE_DIR, "static")
SAMPLES_DIR = os.path.join(STATIC_DIR, "samples")
os.makedirs(SAMPLES_DIR, exist_ok=True)

def get_conn():
    return sqlite3.connect(DB_PATH)

def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL
                )''')
    conn.commit()
    conn.close()

init_db()

# --------- Try importing openpyxl once at startup for clear diagnostics ----------
OPENPYXL_IMPORT_ERROR = None
LOAD_WORKBOOK = None
try:
    from openpyxl import load_workbook  # type: ignore
    LOAD_WORKBOOK = load_workbook
except Exception as e:
    OPENPYXL_IMPORT_ERROR = repr(e)

# ------------------- Auth & Core -------------------
@app.route('/')
def home():
    if session.get("username"):
        return redirect(url_for('claims'))
    return render_template('login.html')

@app.route('/s3display')
def s3display():
    return render_template('s3display.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')

    if not username or not password:
        return render_template('login.html', error="Username and password are required"), 400

    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT password FROM users WHERE username = ?", (username,))
    row = c.fetchone()
    conn.close()

    if row is None:
        return render_template('login.html', error="Invalid username or password"), 400

    stored_hash = row[0]
    if check_password_hash(stored_hash, password):
        session['username'] = username
        return redirect(url_for('claims'))
    else:
        return render_template('login.html', error="Invalid username or password"), 400

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))

# ------------------- UI pages -------------------
@app.route('/claims')
def claims():
    return render_template('claims.html')

# ✅ FIXED: use real Flask dynamic segment
@app.route('/review/<claim_id>')
def review_claim(claim_id):
    return render_template('review.html', claim_id=claim_id)

# ------------------- Excel Preview (XLSM) + Download -------------------

from html import escape
from openpyxl.utils import get_column_letter

def _argb_to_css(argb):
    """Convert openpyxl ARGB like 'FF00FF00' -> '#00FF00'."""
    if not argb:
        return None
    argb = str(argb).replace("0x", "").upper()
    if len(argb) == 8:
        return f"#{argb[2:]}"  # remove alpha
    if len(argb) == 6:
        return f"#{argb}"
    return None

def _border_css(border):
    """Map openpyxl border styles to CSS for each side."""
    if border is None:
        return ""

    style_map = {
        None:   "1px solid #e0e0e0",
        "thin": "1px solid #2b2b2b",
        "medium": "2px solid #2b2b2b",
        "thick": "3px solid #2b2b2b",
        "dashed": "1px dashed #2b2b2b",
        "dotted": "1px dotted #2b2b2b",
        "double": "3px double #2b2b2b",
    }

    def side(side_obj):
        if side_obj is None:
            return style_map[None]
        return style_map.get(side_obj.style, style_map[None])

    return (
        f"border-top:{side(border.top)};"
        f"border-right:{side(border.right)};"
        f"border-bottom:{side(border.bottom)};"
        f"border-left:{side(border.left)};"
    )

def _cell_css(cell):
    """Convert openpyxl cell formatting into inline CSS."""
    styles = []

    # Fill (background)
    fill = cell.fill
    if fill and fill.patternType == "solid":
        fg = _argb_to_css(getattr(fill.fgColor, "rgb", None))
        if fg:
            styles.append(f"background:{fg};")

    # Font
    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight:700;")
        if font.italic:
            styles.append("font-style:italic;")
        if font.size:
            styles.append(f"font-size:{int(font.size)}px;")
        if font.color and getattr(font.color, "rgb", None):
            c = _argb_to_css(font.color.rgb)
            if c:
                styles.append(f"color:{c};")

    # Alignment
    al = cell.alignment
    if al:
        if al.horizontal:
            # openpyxl may return 'centerContinuous' etc; safe map:
            h = "center" if "center" in al.horizontal else al.horizontal
            styles.append(f"text-align:{h};")
        if al.vertical:
            styles.append(f"vertical-align:{al.vertical};")
        if al.wrap_text:
            styles.append("white-space:pre-wrap;")

    # Border
    styles.append(_border_css(cell.border))

    return "".join(styles)

def _compute_used_range(ws, max_rows=250, max_cols=80):
    """
    Compute a reasonable preview range.
    Uses ws.calculate_dimension() when possible; falls back to ws.min/max.
    """
    # Default worksheet bounds
    min_row, min_col = ws.min_row, ws.min_column
    max_row, max_col = ws.max_row, ws.max_column

    # Limit to keep preview fast
    max_row = min(max_row, min_row + max_rows - 1)
    max_col = min(max_col, min_col + max_cols - 1)

    return min_row, min_col, max_row, max_col

def ws_to_html(ws, max_rows=250, max_cols=80):
    """
    Convert a worksheet to styled HTML, supporting merged cells and basic formatting.
    """
    min_row, min_col, max_row, max_col = _compute_used_range(ws, max_rows, max_cols)

    # Build merged cell maps
    merged_map = {}   # (r,c) -> (rowspan, colspan)
    skip = set()      # all cells covered by merged area except top-left

    for r in ws.merged_cells.ranges:
        r1, c1, r2, c2 = r.min_row, r.min_col, r.max_row, r.max_col
        # ignore merges outside preview range
        if r2 < min_row or r1 > max_row or c2 < min_col or c1 > max_col:
            continue

        # clamp merged region to preview region
        rr1, cc1 = max(r1, min_row), max(c1, min_col)
        rr2, cc2 = min(r2, max_row), min(c2, max_col)

        merged_map[(rr1, cc1)] = (rr2 - rr1 + 1, cc2 - cc1 + 1)
        for rr in range(rr1, rr2 + 1):
            for cc in range(cc1, cc2 + 1):
                if (rr, cc) != (rr1, cc1):
                    skip.add((rr, cc))

    # Column widths -> approximate pixels
    colgroup = []
    for c in range(min_col, max_col + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        w = getattr(dim, "width", None)
        px = int(w * 7) if w else 90
        colgroup.append(f'<col style="width:{px}px;">')

    # Build HTML
    html = []
    html.append('<table class="excel-table excel-like">')
    html.append("<colgroup>" + "".join(colgroup) + "</colgroup>")
    html.append("<tbody>")

    for r in range(min_row, max_row + 1):
        # Row height
        rd = ws.row_dimensions.get(r)
        h = getattr(rd, "height", None)
        tr_style = f' style="height:{int(h)}px;"' if h else ""
        html.append(f"<tr{tr_style}>")

        for c in range(min_col, max_col + 1):
            if (r, c) in skip:
                continue

            cell = ws.cell(row=r, column=c)

            # merged spans
            rowspan, colspan = merged_map.get((r, c), (1, 1))
            span = ""
            if rowspan > 1:
                span += f' rowspan="{rowspan}"'
            if colspan > 1:
                span += f' colspan="{colspan}"'

            value = "" if cell.value is None else escape(str(cell.value))
            style = _cell_css(cell)

            html.append(f'<td{span} style="{style}">{value}</td>')

        html.append("</tr>")

    html.append("</tbody></table>")
    return "".join(html)

@app.route('/excel-preview/<claim_id>')
def excel_preview(claim_id):
    if LOAD_WORKBOOK is None:
        details = (
            f"Import error: {OPENPYXL_IMPORT_ERROR}\n"
            f"Python: {sys.version}\n"
            f"Executable: {sys.executable}\n"
            f"Sys.path[0]: {sys.path[0]}"
        )
        return render_template(
            'excel_preview.html',
            claim_id=claim_id,
            sheets=[],
            error="openpyxl import failed. See details below.",
            diag=details
        )

    sample_filename = 'dummy.xlsm'
    sample_abs_path = os.path.join(SAMPLES_DIR, sample_filename)

    if not os.path.exists(sample_abs_path):
        return render_template(
            'excel_preview.html',
            claim_id=claim_id,
            sheets=[],
            error="Dummy XLSM not found at static/samples/dummy.xlsm",
            diag=None
        )

    wb = None
    try:
        # ✅ remove read_only=True to preserve styles/merges/dimensions
        wb = LOAD_WORKBOOK(sample_abs_path, data_only=True, keep_vba=True)

        sheets = []
        for ws in wb.worksheets:
            html_table = ws_to_html(ws, max_rows=250, max_cols=80)
            sheets.append({"name": ws.title, "table_html": html_table})

    except Exception as e:
        return render_template(
            'excel_preview.html',
            claim_id=claim_id,
            sheets=[],
            error=f"Failed to read dummy.xlsm: {e}",
            diag=f"Python: {sys.version}\nExecutable: {sys.executable}"
        )
    finally:
        if wb:
            wb.close()

    return render_template('excel_preview.html', claim_id=claim_id, sheets=sheets, error=None, diag=None)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from html import escape

def argb_to_css(argb):
    """Convert openpyxl ARGB like 'FF00FF00' to '#00FF00'."""
    if not argb:
        return None
    argb = argb.replace("0x", "").upper()
    if len(argb) == 8:
        return f"#{argb[2:]}"  # drop alpha
    if len(argb) == 6:
        return f"#{argb}"
    return None

def border_css(border):
    """Map openpyxl border styles to CSS."""
    if border is None:
        return ""
    # Excel-like thickness mapping
    style_map = {
        None: "1px solid #e0e0e0",
        "thin": "1px solid #2b2b2b",
        "medium": "2px solid #2b2b2b",
        "thick": "3px solid #2b2b2b",
        "dashed": "1px dashed #2b2b2b",
        "dotted": "1px dotted #2b2b2b",
        "double": "3px double #2b2b2b",
    }

    def side(side_obj):
        if side_obj is None:
            return style_map[None]
        return style_map.get(side_obj.style, style_map[None])

    return (
        f"border-top:{side(border.top)};"
        f"border-right:{side(border.right)};"
        f"border-bottom:{side(border.bottom)};"
        f"border-left:{side(border.left)};"
    )

def cell_style(cell):
    """Extract fill/font/alignment/border into inline CSS."""
    styles = []

    # Fill
    fill = cell.fill
    if fill and fill.patternType == "solid":
        fg = argb_to_css(getattr(fill.fgColor, "rgb", None))
        if fg:
            styles.append(f"background:{fg};")

    # Font
    font = cell.font
    if font:
        if font.bold:
            styles.append("font-weight:700;")
        if font.italic:
            styles.append("font-style:italic;")
        if font.size:
            styles.append(f"font-size:{font.size}px;")
        if font.color and getattr(font.color, "rgb", None):
            c = argb_to_css(font.color.rgb)
            if c:
                styles.append(f"color:{c};")

    # Alignment
    al = cell.alignment
    if al:
        if al.horizontal:
            styles.append(f"text-align:{al.horizontal};")
        if al.vertical:
            styles.append(f"vertical-align:{al.vertical};")
        if al.wrap_text:
            styles.append("white-space:pre-wrap;")

    # Border
    styles.append(border_css(cell.border))

    return "".join(styles)

def sheet_to_html(ws, max_rows=80, max_cols=30):
    """
    Convert worksheet to HTML table with Excel-like styles.
    Adjust max_rows/max_cols depending on preview requirements.
    """
    # Detect used range
    min_row = ws.min_row
    min_col = ws.min_column
    max_row_used = min(ws.max_row, min_row + max_rows - 1)
    max_col_used = min(ws.max_column, min_col + max_cols - 1)

    # Track merged cells
    merged_map = {}  # (r,c) -> (rowspan, colspan) and skip others
    skip_cells = set()

    for merged in ws.merged_cells.ranges:
        r1, c1, r2, c2 = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        merged_map[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                if (rr, cc) != (r1, c1):
                    skip_cells.add((rr, cc))

    # Column widths
    colgroup = []
    for c in range(min_col, max_col_used + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        width = getattr(dim, "width", None)
        # Excel width ~ character count; approximate px
        px = int(width * 7) if width else 90
        colgroup.append(f'<col style="width:{px}px;">')

    # Build HTML
    html = []
    html.append('<table class="excel-table excel-like">')
    html.append("<colgroup>" + "".join(colgroup) + "</colgroup>")
    html.append("<tbody>")

    for r in range(min_row, max_row_used + 1):
        # Row height
        rd = ws.row_dimensions.get(r)
        height = getattr(rd, "height", None)
        tr_style = f'style="height:{int(height)}px;"' if height else ""

        html.append(f"<tr {tr_style}>")

        for c in range(min_col, max_col_used + 1):
            if (r, c) in skip_cells:
                continue

            cell = ws.cell(row=r, column=c)

            # Handle merged cell
            rowspan, colspan = merged_map.get((r, c), (1, 1))
            span_attr = ""
            if rowspan > 1:
                span_attr += f' rowspan="{rowspan}"'
            if colspan > 1:
                span_attr += f' colspan="{colspan}"'

            val = "" if cell.value is None else escape(str(cell.value))
            style = cell_style(cell)

            html.append(f'<td{span_attr} style="{style}">{val}</td>')

        html.append("</tr>")

    html.append("</tbody></table>")
    return "".join(html)

def xlsm_to_preview(path):
    wb = load_workbook(path, keep_vba=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "table_html": sheet_to_html(ws)
        })
    return sheets


@app.route('/excel/download')
def excel_download():
    """
    Downloads the dummy.xlsm file from static/samples/.
    """
    filename = 'dummy.xlsm'
    full_path = os.path.join(SAMPLES_DIR, filename)
    if not os.path.exists(full_path):
        abort(404, description="Dummy XLSM not found at static/samples/dummy.xlsm")
    return send_from_directory(SAMPLES_DIR, filename, as_attachment=True)

# -------- Optional: quick diagnostics route --------
@app.route('/diag')
def diag():
    return {
        "python_version": sys.version,
        "executable": sys.executable,
        "openpyxl_import_error": OPENPYXL_IMPORT_ERROR,
        "cwd": os.getcwd(),
        "base_dir": BASE_DIR,
        "static_dir": STATIC_DIR,
        "samples_dir_exists": os.path.isdir(SAMPLES_DIR),
        "dummy_exists": os.path.exists(os.path.join(SAMPLES_DIR, "dummy.xlsm")),
        "sys_path_0": sys.path[0],
    }

if __name__ == '__main__':
    app.run(debug=True)
